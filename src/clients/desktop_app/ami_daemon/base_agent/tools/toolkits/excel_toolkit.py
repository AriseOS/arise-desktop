"""
ExcelToolkit - Excel spreadsheet operations for document management.

Based on Eigent's ExcelToolkit implementation which wraps CAMEL's ExcelToolkit.
Provides Excel file creation and manipulation capabilities.

References:
- Eigent: third-party/eigent/backend/app/utils/toolkit/excel_toolkit.py
- CAMEL: camel.toolkits.ExcelToolkit
"""

import logging
from pathlib import Path
from typing import Any, List, Optional, Union

from .base_toolkit import BaseToolkit, FunctionTool
from ...events import listen_toolkit
from ...workspace import get_working_directory

logger = logging.getLogger(__name__)


class ExcelToolkit(BaseToolkit):
    """A toolkit for Excel spreadsheet operations.

    Provides Excel file capabilities:
    - Extract and analyze content from Excel files (.xlsx, .xls, .csv)
    - Create new Excel workbooks with multiple sheets
    - Read, write, and update cells
    - Row and column manipulation
    - Export to CSV format

    Based on Eigent's implementation which wraps CAMEL's ExcelToolkit.
    """

    agent_name: str = "document_agent"

    def __init__(
        self,
        working_directory: Optional[str] = None,
        timeout: Optional[float] = 60.0,
    ) -> None:
        """Initialize the ExcelToolkit.

        Args:
            working_directory: Directory for file operations.
                If not provided, uses task workspace from WorkingDirectoryManager.
            timeout: Operation timeout in seconds.
        """
        super().__init__(timeout=timeout)

        # Determine working directory - fail if not provided and no workspace manager
        if working_directory:
            self._working_directory = Path(working_directory)
        else:
            self._working_directory = Path(get_working_directory())

        # Ensure directory exists
        self._working_directory.mkdir(parents=True, exist_ok=True)

        logger.info(f"ExcelToolkit initialized in {self._working_directory}")

    def _resolve_filepath(self, filename: str) -> Path:
        """Resolve a filename to an absolute path."""
        path = Path(filename)
        if path.is_absolute():
            return path
        return self._working_directory / filename

    @listen_toolkit(
        inputs=lambda self, filename, **kw: f"Creating Excel workbook: {filename}",
        return_msg=lambda r: r[:200] if isinstance(r, str) else str(r)
    )
    def create_workbook(
        self,
        filename: str,
        sheets: Optional[List[str]] = None,
    ) -> str:
        """Create a new Excel workbook.

        Args:
            filename: Output filename (will add .xlsx if not present).
            sheets: Optional list of sheet names to create.
                If not provided, creates a single sheet named "Sheet1".

        Returns:
            Success message with file path, or error message.
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            return "Error: openpyxl package required. Install with: pip install openpyxl"

        # Ensure filename has .xlsx extension
        if not filename.lower().endswith(('.xlsx', '.xls')):
            filename += ".xlsx"

        filepath = self._resolve_filepath(filename)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Creating Excel workbook: {filepath}")

        try:
            wb = Workbook()

            # Set up sheets
            if sheets:
                # Remove default sheet and create requested sheets
                default_sheet = wb.active
                for i, sheet_name in enumerate(sheets):
                    if i == 0:
                        default_sheet.title = sheet_name
                    else:
                        wb.create_sheet(title=sheet_name)
            else:
                wb.active.title = "Sheet1"

            wb.save(str(filepath))
            logger.info(f"Excel workbook created: {filepath}")
            return f"Excel workbook successfully created: {filepath}"

        except Exception as e:
            error_msg = f"Error creating workbook: {str(e)}"
            logger.error(error_msg)
            return error_msg

    @listen_toolkit(
        inputs=lambda self, filepath, **kw: f"Reading Excel file: {filepath}",
        return_msg=lambda r: f"Read {len(r)} rows" if isinstance(r, list) else str(r)[:200]
    )
    def read_excel(
        self,
        filepath: str,
        sheet_name: Optional[str] = None,
        include_header: bool = True,
    ) -> Union[List[List[Any]], str]:
        """Read data from an Excel file.

        Args:
            filepath: Path to the Excel file.
            sheet_name: Name of sheet to read. Defaults to first sheet.
            include_header: Whether to include header row in output.

        Returns:
            List of rows (each row is a list of cell values), or error message.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl package required. Install with: pip install openpyxl"

        path = self._resolve_filepath(filepath)

        if not path.exists():
            return f"Error: File not found: {path}"

        logger.info(f"Reading Excel file: {path}")

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)

            # Get sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Read data
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))

            wb.close()

            if not include_header and data:
                data = data[1:]

            logger.info(f"Read {len(data)} rows from {path}")
            return data

        except Exception as e:
            error_msg = f"Error reading Excel file: {str(e)}"
            logger.error(error_msg)
            return error_msg

    @listen_toolkit(
        inputs=lambda self, filepath, data, **kw: f"Writing to Excel: {filepath}",
        return_msg=lambda r: r[:200] if isinstance(r, str) else str(r)
    )
    def write_excel(
        self,
        filepath: str,
        data: List[List[Any]],
        sheet_name: Optional[str] = None,
        headers: Optional[List[str]] = None,
    ) -> str:
        """Write data to an Excel file.

        Args:
            filepath: Path to the Excel file (creates if not exists).
            data: List of rows to write.
            sheet_name: Name of sheet to write to.
            headers: Optional list of column headers.

        Returns:
            Success message or error message.
        """
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            return "Error: openpyxl package required. Install with: pip install openpyxl"

        path = self._resolve_filepath(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Writing to Excel file: {path}")

        try:
            # Load existing or create new workbook
            if path.exists():
                wb = load_workbook(str(path))
            else:
                wb = Workbook()

            # Get or create sheet
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb.active

            # Clear existing data
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

            # Write headers
            start_row = 1
            if headers:
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
                start_row = 2

            # Write data
            for row_idx, row_data in enumerate(data, start=start_row):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(str(path))
            logger.info(f"Successfully wrote {len(data)} rows to {path}")
            return f"Successfully wrote {len(data)} rows to {path}"

        except Exception as e:
            error_msg = f"Error writing to Excel file: {str(e)}"
            logger.error(error_msg)
            return error_msg

    @listen_toolkit(
        inputs=lambda self, filepath, row, col, **kw: f"Updating cell ({row}, {col}) in {filepath}",
        return_msg=lambda r: r[:200] if isinstance(r, str) else str(r)
    )
    def update_cell(
        self,
        filepath: str,
        row: int,
        col: int,
        value: Any,
        sheet_name: Optional[str] = None,
    ) -> str:
        """Update a specific cell in an Excel file.

        Args:
            filepath: Path to the Excel file.
            row: Row number (1-indexed).
            col: Column number (1-indexed).
            value: Value to set.
            sheet_name: Name of sheet.

        Returns:
            Success message or error message.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl package required. Install with: pip install openpyxl"

        path = self._resolve_filepath(filepath)

        if not path.exists():
            return f"Error: File not found: {path}"

        try:
            wb = load_workbook(str(path))

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return f"Error: Sheet '{sheet_name}' not found"
                ws = wb[sheet_name]
            else:
                ws = wb.active

            ws.cell(row=row, column=col, value=value)
            wb.save(str(path))

            logger.info(f"Updated cell ({row}, {col}) in {path}")
            return f"Successfully updated cell ({row}, {col}) to '{value}'"

        except Exception as e:
            error_msg = f"Error updating cell: {str(e)}"
            logger.error(error_msg)
            return error_msg

    @listen_toolkit(
        inputs=lambda self, filepath, **kw: f"Getting sheet names from {filepath}",
        return_msg=lambda r: str(r)
    )
    def get_sheet_names(self, filepath: str) -> Union[List[str], str]:
        """Get all sheet names in an Excel file.

        Args:
            filepath: Path to the Excel file.

        Returns:
            List of sheet names, or error message.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl package required. Install with: pip install openpyxl"

        path = self._resolve_filepath(filepath)

        if not path.exists():
            return f"Error: File not found: {path}"

        try:
            wb = load_workbook(str(path), read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets

        except Exception as e:
            return f"Error getting sheet names: {str(e)}"

    @listen_toolkit(
        inputs=lambda self, filepath, output_path, **kw: f"Exporting {filepath} to CSV",
        return_msg=lambda r: r[:200] if isinstance(r, str) else str(r)
    )
    def export_to_csv(
        self,
        filepath: str,
        output_path: Optional[str] = None,
        sheet_name: Optional[str] = None,
    ) -> str:
        """Export an Excel sheet to CSV format.

        Args:
            filepath: Path to the Excel file.
            output_path: Output CSV path. Defaults to same name with .csv extension.
            sheet_name: Sheet to export. Defaults to first sheet.

        Returns:
            Success message with output path, or error message.
        """
        import csv

        # Read Excel data
        data = self.read_excel(filepath, sheet_name=sheet_name)
        if isinstance(data, str):
            return data  # Error message

        # Determine output path
        if output_path:
            out_path = self._resolve_filepath(output_path)
        else:
            in_path = self._resolve_filepath(filepath)
            out_path = in_path.with_suffix('.csv')

        out_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            with open(out_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in data:
                    writer.writerow(row)

            logger.info(f"Exported to CSV: {out_path}")
            return f"Successfully exported to CSV: {out_path}"

        except Exception as e:
            error_msg = f"Error exporting to CSV: {str(e)}"
            logger.error(error_msg)
            return error_msg

    @property
    def working_directory(self) -> Path:
        """Get the current working directory."""
        return self._working_directory

    def get_tools(self) -> List[FunctionTool]:
        """Return a list of FunctionTool objects for this toolkit.

        Returns:
            List of FunctionTool objects.
        """
        return [
            FunctionTool(self.create_workbook),
            FunctionTool(self.read_excel),
            FunctionTool(self.write_excel),
            FunctionTool(self.update_cell),
            FunctionTool(self.get_sheet_names),
            FunctionTool(self.export_to_csv),
        ]

    @classmethod
    def toolkit_name(cls) -> str:
        """Return the name of this toolkit."""
        return "Excel"
